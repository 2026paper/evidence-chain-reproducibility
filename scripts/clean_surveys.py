from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PROJECT_ROOT.parent
SURVEY_ROOT = WORKSPACE_ROOT / "数据与问卷源文件"
SECOND_REVIEW_ROOT = SURVEY_ROOT / "专家复核问卷"
OUTPUT_ROOT = SURVEY_ROOT / "清洗后的数据"
TMP_ROOT = PROJECT_ROOT / "tmp" / "spreadsheets"
ANALYSIS_ROOT = PROJECT_ROOT / "analysis"

SIMILARITY_THRESHOLD = 90.0
SENSITIVITY_THRESHOLDS = (80.0, 85.0, 90.0, 95.0, 100.0)
DIMENSIONS = (
    "事实准确性",
    "概念完整性",
    "语言清晰度",
    "任务符合度",
    "误解处理质量",
    "误导风险",
)

PUBLIC_FILES = {"A.xlsx", "B.xlsx", "C.xlsx"}
FIRST_EXPERT_FILES = {"化学.xlsx", "地理.xlsx", "大气科学.xlsx", "物理.xlsx", "生物.xlsx"}
DOMAIN_BY_FILE = {
    "化学.xlsx": "化学",
    "地理.xlsx": "地理",
    "大气科学.xlsx": "大气科学",
    "物理.xlsx": "物理",
    "生物.xlsx": "生物",
}
DOMAIN_BY_CASE_PREFIX = {
    "A": "大气科学",
    "B": "生物",
    "C": "化学",
    "G": "地理",
    "P": "物理",
}

SHEET_NAME_BY_FILE = {
    "A.xlsx": "首轮公众_A",
    "B.xlsx": "首轮公众_B",
    "C.xlsx": "首轮公众_C",
    "分歧.xlsx": "分歧验证",
    "化学.xlsx": "首轮专家_化学",
    "地理.xlsx": "首轮专家_地理",
    "大气科学.xlsx": "首轮专家_大气",
    "差异.xlsx": "差异对比",
    "物理.xlsx": "首轮专家_物理",
    "生物.xlsx": "首轮专家_生物",
    "质性.xlsx": "质性反馈",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def pseudonym(source: str, raw_id: Any) -> str:
    value = f"{source}|{raw_id}".encode("utf-8")
    return hashlib.sha256(value).hexdigest()[:12]


def privacy_safe_column_indices(headers: list[Any]) -> list[int]:
    """Return questionnaire columns that are safe for the cleaned workbook.

    Platform identifiers, timestamps, location, device/browser metadata and common
    direct identifiers remain only in the untouched raw exports.  The cleaned
    workbook receives a deterministic pseudonym in their place.
    """
    exact = {"ip", "ip\u5730\u5740"}
    fragments = (
        "\u4f5c\u7b54id", "\u7528\u6237id", "\u53d1\u5e03id",
        "\u5f00\u59cb\u65f6\u95f4", "\u7ed3\u675f\u65f6\u95f4", "\u4f5c\u7b54\u603b\u65f6\u957f", "\u7b54\u9898\u65f6\u957f",
        "\u4f5c\u7b54\u6e20\u9053", "\u95ee\u5377\u53d1\u5e03\u540d\u79f0",
        "\u7ecf\u5ea6", "\u7eac\u5ea6", "\u7701\u4efd", "\u57ce\u5e02", "\u5730\u533a", "\u6240\u5728\u5730",
        "\u8bbe\u5907", "\u64cd\u4f5c\u7cfb\u7edf", "\u6d4f\u89c8\u5668", "\u5c4f\u5e55\u5206\u8fa8\u7387",
        "\u59d3\u540d", "\u624b\u673a", "\u7535\u8bdd", "\u90ae\u7bb1", "\u5fae\u4fe1", "qq", "\u8eab\u4efd\u8bc1",
    )
    keep: list[int] = []
    for index, header in enumerate(headers):
        normalized = str(header or "").strip().casefold().replace(" ", "")
        if normalized in exact or any(fragment in normalized for fragment in fragments):
            continue
        keep.append(index)
    return keep


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def normalized_manhattan_similarity(left: Iterable[Any], right: Iterable[Any]) -> dict[str, float | None]:
    a = [as_float(value) for value in left]
    b = [as_float(value) for value in right]
    if len(a) != len(b) or not a or any(value is None for value in a + b):
        return {"similarity": None, "mae": None, "exact_agreement": None}
    diffs = [abs(float(x) - float(y)) for x, y in zip(a, b)]
    similarity = 100.0 * (1.0 - sum(diffs) / (4.0 * len(diffs)))
    exact = 100.0 * sum(diff == 0 for diff in diffs) / len(diffs)
    return {
        "similarity": round(similarity, 6),
        "mae": round(sum(diffs) / len(diffs), 6),
        "exact_agreement": round(exact, 6),
    }


@dataclass
class WorkbookData:
    path: Path
    headers: list[Any]
    code_row: list[Any] | None
    rows: list[list[Any]]
    excel_row_numbers: list[int]


def read_credamo_workbook(path: Path) -> WorkbookData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    # Credamo exports incorrectly declare dimension A1 although the sheet contains
    # hundreds of columns. Reset the cached dimension before iterating.
    worksheet.reset_dimensions()
    values = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()
    if not values:
        raise ValueError(f"Empty workbook: {path}")

    headers = values[0]
    code_row = None
    start = 1
    if len(values) > 1 and values[1] and str(values[1][0]) == str(headers[0]):
        code_row = values[1]
        start = 2

    rows: list[list[Any]] = []
    excel_row_numbers: list[int] = []
    for zero_index, row in enumerate(values[start:], start=start):
        if not any(value is not None and value != "" for value in row):
            continue
        if len(row) < len(headers):
            row = row + [None] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[: len(headers)]
        rows.append(row)
        excel_row_numbers.append(zero_index + 1)
    return WorkbookData(path, headers, code_row, rows, excel_row_numbers)


def source_kind(path: Path, headers: list[Any]) -> tuple[str, str, str]:
    if path.parent == SECOND_REVIEW_ROOT:
        prefix = second_review_prefix(headers)
        domain = DOMAIN_BY_CASE_PREFIX[prefix]
        return "二次专家复核", domain, f"二次专家_{domain.replace('大气科学', '大气')}"
    if path.name in FIRST_EXPERT_FILES:
        return "首轮专家复核", DOMAIN_BY_FILE[path.name], SHEET_NAME_BY_FILE[path.name]
    if path.name in PUBLIC_FILES:
        return "首轮公众测试", "跨学科", SHEET_NAME_BY_FILE[path.name]
    if path.name in SHEET_NAME_BY_FILE:
        return "辅助人工问卷", "跨学科", SHEET_NAME_BY_FILE[path.name]
    raise ValueError(f"Unclassified source: {path}")


def second_review_prefix(headers: list[Any]) -> str:
    for value in headers:
        match = re.search(r"案例\s+([A-Z])\d{2}", str(value or ""))
        if match:
            return match.group(1)
    raise ValueError("No second-review case prefix found")


def code_index(code_row: list[Any] | None) -> dict[str, int]:
    if code_row is None:
        return {}
    return {str(value): index for index, value in enumerate(code_row) if value not in (None, "")}


def exact_attention_result(got: list[Any], expected: list[Any]) -> tuple[bool, int]:
    if len(got) != len(expected):
        return False, max(len(got), len(expected))
    failures = 0
    for observed, target in zip(got, expected):
        if isinstance(target, (int, float)):
            observed_number = as_float(observed)
            ok = observed_number is not None and observed_number == float(target)
        else:
            ok = str(observed or "").strip() == str(target).strip()
        failures += int(not ok)
    return failures == 0, failures


def attention_spec(data: WorkbookData, wave: str) -> tuple[list[int], list[Any], str]:
    headers = [str(value or "") for value in data.headers]
    if wave == "首轮专家复核":
        q6 = [i for i, text in enumerate(headers) if "评分校准样例" in text]
        q19 = [i for i, text in enumerate(headers) if "事实准确性为5分" in text]
        q45 = [i for i, text in enumerate(headers) if "各维度均为4分" in text]
        indices = q6 + q19 + q45
        expected = [3] * 6 + [5, 2, 2, 2, 2, 2] + [4] * 6
        if len(indices) != 18:
            raise ValueError(f"Expected 18 keyed attention cells in {data.path}, got {len(indices)}")
        return indices, expected, "Q6 all 3; Q19 = [5,2,2,2,2,2]; Q45 all 4"

    if wave == "二次专家复核":
        indices = [i for i, text in enumerate(headers) if "以下题不评价任何材料" in text]
        if len(indices) != 1:
            raise ValueError(f"Expected one risk-direction attention item in {data.path}")
        return indices, ["分数越高，普通读者越可能被内容误导"], "Risk direction"

    if data.path.name in PUBLIC_FILES:
        indices = [i for i, text in enumerate(headers) if "为了确认您认真阅读" in text]
        return indices, ["比较同意"], "Public attention item"

    if data.path.name == "分歧.xlsx":
        indices = [i for i, text in enumerate(headers) if "为了确认您认真阅读说明" in text]
        return indices, ["评委身份已匿名化"], "Judge-identity attention item"

    return [], [], "Not applicable"


def first_expert_repeat_indices(data: WorkbookData, first: str, second: str) -> tuple[list[int], list[int]]:
    mapping = code_index(data.code_row)
    left = [mapping.get(f"{first}_{dimension}") for dimension in range(1, 7)]
    right = [mapping.get(f"{second}_{dimension}") for dimension in range(1, 7)]
    if any(index is None for index in left + right):
        raise ValueError(f"Missing repeat mapping {first}/{second} in {data.path}")
    return [int(index) for index in left], [int(index) for index in right]


def second_expert_repeat_indices(data: WorkbookData, case_number: str) -> list[int]:
    prefix = second_review_prefix(data.headers)
    case_id = f"{prefix}{case_number}"
    indices: list[int] = []
    for index, value in enumerate(data.headers):
        text = str(value or "")
        if case_id in text and any(text.endswith(f"-{dimension}") for dimension in DIMENSIONS):
            indices.append(index)
    if len(indices) != 6:
        raise ValueError(f"Expected six score columns for {case_id} in {data.path}, got {len(indices)}")
    return indices


def repeat_specs(data: WorkbookData, wave: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    primary: list[dict[str, Any]] = []
    diagnostic: list[dict[str, Any]] = []
    if wave == "首轮专家复核":
        left, right = first_expert_repeat_indices(data, "Q39", "Q46")
        primary.append({"name": "Q39/Q46", "left": left, "right": right})
        diag_left, diag_right = first_expert_repeat_indices(data, "Q12", "Q32")
        diagnostic.append({"name": "Q12/Q32", "left": diag_left, "right": diag_right})
    elif wave == "二次专家复核":
        primary.append(
            {
                "name": "Case05/Case07",
                "left": second_expert_repeat_indices(data, "05"),
                "right": second_expert_repeat_indices(data, "07"),
            }
        )
    return primary, diagnostic


def row_repeat_metrics(row: list[Any], spec: dict[str, Any]) -> dict[str, Any]:
    left = [row[index] for index in spec["left"]]
    right = [row[index] for index in spec["right"]]
    metrics = normalized_manhattan_similarity(left, right)
    return {"name": spec["name"], **metrics}


def summarize_numbers(values: list[float]) -> tuple[float | None, float | None, float | None]:
    if not values:
        return None, None, None
    return round(min(values), 6), round(median(values), 6), round(max(values), 6)


def infer_second_task(text: str) -> str:
    if "科学选择题" in text or "【题干】" in text:
        return "Task C"
    if "情境科普卡片" in text or "标题：" in text:
        return "Task B"
    return "Task A"


def substantive_first_questions() -> list[int]:
    return list(range(7, 19)) + list(range(20, 32)) + list(range(33, 45))


def first_expert_long_rows(
    data: WorkbookData,
    retained_rows: list[list[Any]],
    retained_ids: list[str],
    domain: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping = code_index(data.code_row)
    ratings: list[dict[str, Any]] = []
    dictionary: dict[tuple[str, str], dict[str, Any]] = {}
    item_meta: dict[int, tuple[str, str, str]] = {}
    for qnum in substantive_first_questions():
        task = "Task A" if qnum <= 18 else ("Task B" if qnum <= 31 else "Task C")
        first_index = mapping[f"Q{qnum}_1"]
        header = str(data.headers[first_index] or "")
        item_text = header.rsplit("-", 1)[0]
        item_hash = hashlib.sha256(item_text.encode("utf-8")).hexdigest()[:16]
        item_meta[qnum] = (task, item_text, item_hash)
        dictionary[(f"Q{qnum}", item_hash)] = {
            "wave": "首轮专家复核",
            "domain": domain,
            "task": task,
            "item_local_id": f"Q{qnum}",
            "item_text_hash": item_hash,
            "item_text": item_text,
        }

    for row, participant_id in zip(retained_rows, retained_ids):
        for qnum in substantive_first_questions():
            task, _item_text, item_hash = item_meta[qnum]
            for dimension_number, dimension in enumerate(DIMENSIONS, start=1):
                value = as_float(row[mapping[f"Q{qnum}_{dimension_number}"]])
                if value is None:
                    continue
                ratings.append(
                    {
                        "wave": "首轮专家复核",
                        "domain": domain,
                        "task": task,
                        "source_file": data.path.name,
                        "participant_id": participant_id,
                        "item_local_id": f"Q{qnum}",
                        "item_text_hash": item_hash,
                        "dimension": dimension,
                        "score_raw": value,
                        "score_quality_aligned": 6.0 - value if dimension == "误导风险" else value,
                    }
                )
    return ratings, list(dictionary.values())


def second_case_score_columns(data: WorkbookData, case_id: str) -> list[int]:
    columns: list[int] = []
    for index, value in enumerate(data.headers):
        text = str(value or "")
        if case_id in text and any(text.endswith(f"-{dimension}") for dimension in DIMENSIONS):
            columns.append(index)
    return columns


def second_case_text(data: WorkbookData, case_id: str) -> str:
    candidates = [str(value or "") for value in data.headers if case_id in str(value or "")]
    if not candidates:
        raise ValueError(f"No text for {case_id}")
    text = max(candidates, key=len)
    if any(text.endswith(f"-{dimension}") for dimension in DIMENSIONS):
        text = text.rsplit("-", 1)[0]
    return text


def second_expert_long_rows(
    data: WorkbookData,
    retained_rows: list[list[Any]],
    retained_ids: list[str],
    domain: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    prefix = second_review_prefix(data.headers)
    ratings: list[dict[str, Any]] = []
    dictionary: list[dict[str, Any]] = []
    for case_number in range(1, 7):
        case_id = f"{prefix}{case_number:02d}"
        columns = second_case_score_columns(data, case_id)
        if len(columns) != 6:
            raise ValueError(f"Expected six columns for {case_id}, got {len(columns)}")
        item_text = second_case_text(data, case_id)
        item_hash = hashlib.sha256(item_text.encode("utf-8")).hexdigest()[:16]
        task = infer_second_task(item_text)
        dictionary.append(
            {
                "wave": "二次专家复核",
                "domain": domain,
                "task": task,
                "item_local_id": case_id,
                "item_text_hash": item_hash,
                "item_text": item_text,
            }
        )
        for row, participant_id in zip(retained_rows, retained_ids):
            for column, dimension in zip(columns, DIMENSIONS):
                value = as_float(row[column])
                if value is None:
                    continue
                ratings.append(
                    {
                        "wave": "二次专家复核",
                        "domain": domain,
                        "task": task,
                        "source_file": data.path.name,
                        "participant_id": participant_id,
                        "item_local_id": case_id,
                        "item_text_hash": item_hash,
                        "dimension": dimension,
                        "score_raw": value,
                        "score_quality_aligned": 6.0 - value if dimension == "误导风险" else value,
                    }
                )
    return ratings, dictionary


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def clean_all() -> dict[str, Any]:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    TMP_ROOT.mkdir(parents=True, exist_ok=True)
    ANALYSIS_ROOT.mkdir(parents=True, exist_ok=True)

    source_paths = sorted(SURVEY_ROOT.glob("*.xlsx"), key=lambda p: p.name)
    source_paths += sorted(SECOND_REVIEW_ROOT.glob("*.xlsx"), key=lambda p: p.name)
    if len(source_paths) != 16:
        raise ValueError(f"Expected 16 questionnaire workbooks, found {len(source_paths)}")

    summaries: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    sensitivity: list[dict[str, Any]] = []
    scenarios: list[dict[str, Any]] = []
    workbook_sheets: list[dict[str, Any]] = []
    long_ratings: list[dict[str, Any]] = []
    item_dictionary: list[dict[str, Any]] = []
    repeat_diagnostics: list[dict[str, Any]] = []

    for path in source_paths:
        data = read_credamo_workbook(path)
        wave, domain, sheet_name = source_kind(path, data.headers)
        relative_source = str(path.relative_to(SURVEY_ROOT))
        attention_columns, attention_expected, attention_rule = attention_spec(data, wave)
        primary_specs, diagnostic_specs = repeat_specs(data, wave)

        retained_rows: list[list[Any]] = []
        retained_ids: list[str] = []
        source_audit: list[dict[str, Any]] = []

        for row, excel_row in zip(data.rows, data.excel_row_numbers):
            participant_id = pseudonym(relative_source, row[0] if row else excel_row)
            observed_attention = [row[index] for index in attention_columns]
            if attention_columns:
                attention_pass, attention_failures = exact_attention_result(observed_attention, attention_expected)
            else:
                attention_pass, attention_failures = True, 0

            primary_metrics = [row_repeat_metrics(row, spec) for spec in primary_specs]
            diagnostic_metrics = [row_repeat_metrics(row, spec) for spec in diagnostic_specs]
            repeat_similarities = [metric["similarity"] for metric in primary_metrics]
            repeat_pass = all(
                similarity is not None and float(similarity) >= SIMILARITY_THRESHOLD
                for similarity in repeat_similarities
            ) if primary_specs else True

            reasons: list[str] = []
            if not attention_pass:
                reasons.append("ATTENTION_FAILED")
            if not repeat_pass:
                reasons.append("REPEAT_SIMILARITY_LT_90")
            keep = not reasons

            if keep:
                retained_rows.append(row)
                retained_ids.append(participant_id)

            primary = primary_metrics[0] if primary_metrics else {}
            diagnostic = diagnostic_metrics[0] if diagnostic_metrics else {}
            audit_row = {
                "source_file": relative_source,
                "wave": wave,
                "domain": domain,
                "participant_id": participant_id,
                "excel_row": excel_row,
                "decision": "保留" if keep else "剔除",
                "reason": ";".join(reasons) if reasons else "PASS",
                "attention_applicable": bool(attention_columns),
                "attention_pass": attention_pass if attention_columns else None,
                "attention_fail_cells": attention_failures if attention_columns else None,
                "repeat_applicable": bool(primary_specs),
                "repeat_pair": primary.get("name"),
                "repeat_similarity_pct": primary.get("similarity"),
                "repeat_mae": primary.get("mae"),
                "repeat_exact_agreement_pct": primary.get("exact_agreement"),
                "diagnostic_repeat_pair": diagnostic.get("name"),
                "diagnostic_repeat_similarity_pct": diagnostic.get("similarity"),
            }
            audit.append(audit_row)
            source_audit.append(audit_row)

            for metric in primary_metrics + diagnostic_metrics:
                repeat_diagnostics.append(
                    {
                        "source_file": relative_source,
                        "wave": wave,
                        "domain": domain,
                        "participant_id": participant_id,
                        "pair": metric["name"],
                        "role": "primary_gate" if metric in primary_metrics else "diagnostic_only",
                        "similarity_pct": metric["similarity"],
                        "mae": metric["mae"],
                        "exact_agreement_pct": metric["exact_agreement"],
                        "attention_pass": attention_pass if attention_columns else None,
                    }
                )

        raw_n = len(data.rows)
        kept_n = len(retained_rows)
        attention_fail_n = sum(row["attention_pass"] is False for row in source_audit)
        repeat_fail_n = sum(row["repeat_applicable"] and (row["repeat_similarity_pct"] is None or row["repeat_similarity_pct"] < SIMILARITY_THRESHOLD) for row in source_audit)
        both_fail_n = sum(row["attention_pass"] is False and row["repeat_applicable"] and (row["repeat_similarity_pct"] is None or row["repeat_similarity_pct"] < SIMILARITY_THRESHOLD) for row in source_audit)
        similarities = [float(row["repeat_similarity_pct"]) for row in source_audit if row["repeat_similarity_pct"] is not None]
        similarity_min, similarity_median, similarity_max = summarize_numbers(similarities)
        summaries.append(
            {
                "source_file": relative_source,
                "sheet_name": sheet_name,
                "wave": wave,
                "domain": domain,
                "raw_responses": raw_n,
                "retained": kept_n,
                "excluded": raw_n - kept_n,
                "retention_rate": kept_n / raw_n if raw_n else None,
                "attention_applicable": bool(attention_columns),
                "attention_rule": attention_rule,
                "attention_failed": attention_fail_n,
                "repeat_applicable": bool(primary_specs),
                "repeat_rule": primary_specs[0]["name"] if primary_specs else "Not applicable",
                "repeat_failed": repeat_fail_n,
                "failed_both": both_fail_n,
                "repeat_similarity_min": similarity_min,
                "repeat_similarity_median": similarity_median,
                "repeat_similarity_max": similarity_max,
                "source_sha256": sha256_file(path),
            }
        )

        for threshold in SENSITIVITY_THRESHOLDS:
            threshold_kept = sum(
                (row["attention_pass"] is not False)
                and (
                    not row["repeat_applicable"]
                    or (row["repeat_similarity_pct"] is not None and row["repeat_similarity_pct"] >= threshold)
                )
                for row in source_audit
            )
            sensitivity.append(
                {
                    "source_file": relative_source,
                    "wave": wave,
                    "domain": domain,
                    "scenario": "primary_repeat_gate",
                    "threshold_pct": threshold,
                    "retained": threshold_kept,
                    "raw_responses": raw_n,
                    "retention_rate": threshold_kept / raw_n if raw_n else None,
                }
            )

        if wave == "首轮专家复核":
            strict_kept = 0
            for row in source_audit:
                primary_ok = row["repeat_similarity_pct"] is not None and row["repeat_similarity_pct"] >= SIMILARITY_THRESHOLD
                diagnostic_ok = row["diagnostic_repeat_similarity_pct"] is not None and row["diagnostic_repeat_similarity_pct"] >= SIMILARITY_THRESHOLD
                strict_kept += int(row["attention_pass"] is not False and primary_ok and diagnostic_ok)
            scenarios.append(
                {
                    "source_file": relative_source,
                    "wave": wave,
                    "domain": domain,
                    "scenario": "strict_both_Q12_Q32_and_Q39_Q46",
                    "threshold_pct": SIMILARITY_THRESHOLD,
                    "retained": strict_kept,
                    "raw_responses": raw_n,
                    "retention_rate": strict_kept / raw_n if raw_n else None,
                }
            )

        safe_indices = privacy_safe_column_indices(data.headers)
        workbook_sheets.append(
            {
                "sheet_name": sheet_name[:31],
                "source_file": relative_source,
                "wave": wave,
                "domain": domain,
                "headers": ["\u533f\u540d\u7b54\u5377ID"]
                + [json_safe(data.headers[index]) for index in safe_indices],
                "rows": [
                    [participant_id]
                    + [json_safe(row[index]) for index in safe_indices]
                    for row, participant_id in zip(retained_rows, retained_ids)
                ],
            }
        )

        if wave == "首轮专家复核":
            ratings, dictionary = first_expert_long_rows(data, retained_rows, retained_ids, domain)
            long_ratings.extend(ratings)
            item_dictionary.extend(dictionary)
        elif wave == "二次专家复核":
            ratings, dictionary = second_expert_long_rows(data, retained_rows, retained_ids, domain)
            long_ratings.extend(ratings)
            item_dictionary.extend(dictionary)

    # Validate user-level totals before authoring any workbook.
    total_raw = sum(row["raw_responses"] for row in summaries)
    if total_raw != 438:
        raise ValueError(f"Expected 438 human questionnaire records, found {total_raw}")
    first_expert_kept = sum(row["retained"] for row in summaries if row["wave"] == "首轮专家复核")
    second_expert_kept = sum(row["retained"] for row in summaries if row["wave"] == "二次专家复核")
    if first_expert_kept != 9 or second_expert_kept != 95:
        raise ValueError(
            f"Cleaning total mismatch: expected first=9 and second=95, got first={first_expert_kept}, second={second_expert_kept}"
        )

    payload = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "survey_root": str(SURVEY_ROOT),
        "output_root": str(OUTPUT_ROOT),
        "similarity_threshold_pct": SIMILARITY_THRESHOLD,
        "similarity_formula": "100 * (1 - sum(abs(x_i-y_i)) / (4*k)), for k paired 1-5 ratings",
        "pass_interpretation": "PASS means exclude the response",
        "dimensions": list(DIMENSIONS),
        "total_raw_questionnaire_records": total_raw,
        "total_retained_records": sum(row["retained"] for row in summaries),
        "summaries": summaries,
        "audit": audit,
        "sensitivity": sensitivity,
        "scenarios": scenarios,
        "sheets": workbook_sheets,
    }

    payload_path = TMP_ROOT / "cleaning_payload.json"
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(TMP_ROOT / "cleaning_summary.csv", summaries)
    write_csv(TMP_ROOT / "cleaning_audit.csv", audit)
    write_csv(TMP_ROOT / "cleaning_sensitivity.csv", sensitivity + scenarios)
    write_csv(ANALYSIS_ROOT / "cleaned_human_ratings_long.csv", long_ratings)
    # De-duplicate dictionary rows while preserving first occurrence.
    seen_dictionary: set[tuple[str, str, str]] = set()
    dictionary_unique: list[dict[str, Any]] = []
    for row in item_dictionary:
        key = (row["wave"], row["domain"], row["item_local_id"])
        if key not in seen_dictionary:
            seen_dictionary.add(key)
            dictionary_unique.append(row)
    write_csv(ANALYSIS_ROOT / "questionnaire_item_dictionary.csv", dictionary_unique)
    write_csv(ANALYSIS_ROOT / "repeat_diagnostics.csv", repeat_diagnostics)

    compact_summary = {
        "total_raw": total_raw,
        "total_retained": payload["total_retained_records"],
        "first_expert_raw": sum(row["raw_responses"] for row in summaries if row["wave"] == "首轮专家复核"),
        "first_expert_retained": first_expert_kept,
        "second_expert_raw": sum(row["raw_responses"] for row in summaries if row["wave"] == "二次专家复核"),
        "second_expert_retained": second_expert_kept,
        "audit_rows": len(audit),
        "long_rating_rows": len(long_ratings),
        "dictionary_items": len(dictionary_unique),
        "payload": str(payload_path),
    }
    (TMP_ROOT / "cleaning_run_summary.json").write_text(
        json.dumps(compact_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return compact_summary


if __name__ == "__main__":
    print(json.dumps(clean_all(), ensure_ascii=False, indent=2))
