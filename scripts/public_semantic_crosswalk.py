#!/usr/bin/env python3
"""Build an auditable, privacy-safe crosswalk for the public-reader materials.

This script never reads participant rows.  It uses the frozen, privacy-safe
instrument text, the released case-provenance table, the controlled-pair
definition, and the 810-item corpus.  Candidate retrieval is diagnostic only;
only mappings supported by an explicit provenance bridge are labelled as
one-to-one mappings.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "数据与问卷源文件"
GITHUB = SOURCE / "AI测试数据全集" / "40_GitHub发布修订"
MAIN = GITHUB / "main_study_data"
AUX = GITHUB / "rebuttal_update_20260714" / "auxiliary_surveys"
INSTR = AUX / "instruments"
ANALYSIS = ROOT / "ADMA2026_重构稿" / "analysis"
NOTES = ROOT / "ADMA2026_重构稿" / "notes"

ITEMS_PATH = MAIN / "generated_items_810.csv"
CASE_PATH = MAIN / "case_disagreement_examples.csv"
CONTROLLED_PATH = GITHUB / "diagnostics" / "controlled_ab_pairs_for_llm_judge.json"
CONTROLLED_SCORE_PATH = (
    GITHUB
    / "diagnostics"
    / "official_ab_judge_run_20260521_155532"
    / "official_ab_judge_scores_long_with_metadata.csv"
)

FORM_PATHS = {
    "A": INSTR / "public_reader_A_instrument_zh.txt",
    "B": INSTR / "public_reader_B_instrument_zh.txt",
    "C": INSTR / "public_reader_C_instrument_zh.txt",
}


ABC_META = {
    1: {
        "source_case_id": "C1",
        "core_claim": "忽略空气阻力时，离手后无持续水平推力，物体依惯性保持水平运动。",
        "target_misconception": "物体只要继续运动，就必然仍有同向的力维持运动。",
        "boundary_conditions": "理想化模型：忽略空气阻力；现实情况需加入阻力、旋转等因素。",
    },
    2: {
        "source_case_id": "C2",
        "core_claim": "进化没有预设终点，可遗传差异在当前环境中影响生存和繁殖。",
        "target_misconception": "进化会主动朝更高级、更完美的目标前进。",
        "boundary_conditions": "适应性是环境依赖的，不等于完美或绝对高级。",
    },
    3: {
        "source_case_id": "C3",
        "core_claim": "缓慢的板块运动在地质时间上可累积为显著结果。",
        "target_misconception": "肉眼短期看不见的变化几乎没有影响，或所有地质事件都只能缓慢发生。",
        "boundary_conditions": "类比说明长时标累积，不意味着所有地质事件匀速或缓慢。",
    },
    4: {
        "source_case_id": "C4",
        "core_claim": "植物昼夜都进行细胞呼吸，光合作用需要光。",
        "target_misconception": "植物白天只光合、夜间才呼吸，或植物不呼吸。",
        "boundary_conditions": "净气体交换表现受光合与呼吸的相对速率影响。",
    },
    5: {
        "source_case_id": "C5",
        "core_claim": "进化无固定复杂化方向，简化在特定环境中也可能有利。",
        "target_misconception": "进化必然从低级到高级、从简单到复杂。",
        "boundary_conditions": "评价依据是当前环境中的生存和繁殖，不是普遍的复杂性尺度。",
    },
    6: {
        "source_case_id": "C6",
        "core_claim": "净力直接决定加速度，而非速度大小。",
        "target_misconception": "力越大速度就一定越大，或速度大就一定有大净力。",
        "boundary_conditions": "匀速直线运动时净力为零，但各个作用力可以平衡而非全部不存在。",
    },
}


CONTROLLED_CONCEPT = {
    "A01": "C04",
    "A02": "C02",
    "A03": "C14",
    "A04": "C18",
    "A05": "C20",
    "A06": "C26",
    "A07": None,
    "A08": "C25",
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize(text: Any) -> str:
    text = "" if text is None else str(text)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[\u3000，。：；！？“”‘’、,.:;!?()\[\]{}<>/《》【】—–_-]", "", text)
    return text.lower()


def normalize_without_format_markers(text: Any) -> str:
    return normalize(re.sub(r"【[^】]*】", "", "" if text is None else str(text)))


def flatten_json_text(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, dict):
        out: list[str] = []
        for v in value.values():
            out.extend(flatten_json_text(v))
        return out
    if isinstance(value, list):
        out = []
        for v in value:
            out.extend(flatten_json_text(v))
        return out
    return [str(value)]


def item_document(row: pd.Series) -> str:
    pieces = [
        str(row.get("concept", "")),
        str(row.get("output", "")),
        str(row.get("text", "")),
    ]
    raw = row.get("content_json", "")
    try:
        parsed = json.loads(raw) if isinstance(raw, str) and raw.strip() else {}
        pieces.extend(flatten_json_text(parsed))
    except json.JSONDecodeError:
        pieces.append(str(raw))
    return normalize(" ".join(pieces))


def parse_public_materials(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8")
    blocks = re.findall(
        r"Q(1[3-8])\s*\n(\u9605\u8bfb\u6750\u6599(\d+)\uff1a.*?)(?=\nQ(?:1[4-9])\b)",
        text,
        flags=re.S,
    )
    result = []
    for qcode, block, material_no in blocks:
        if "\u8bf7\u60a8\u9605\u8bfb\u540e\u5224\u65ad\uff1a" not in block:
            raise RuntimeError(f"Cannot split material/question in {path.name} {qcode}")
        exposure, tail = block.split("\u8bf7\u60a8\u9605\u8bfb\u540e\u5224\u65ad\uff1a", 1)
        question = tail.split("[\u5355\u9009]", 1)[0].strip()
        result.append(
            {
                "qcode": f"Q{qcode}",
                "material_no": int(material_no),
                "exposure_text": exposure.strip(),
                "comprehension_question": question,
            }
        )
    if len(result) != 6:
        raise RuntimeError(f"Expected 6 materials in {path}, found {len(result)}")
    return sorted(result, key=lambda x: x["material_no"])


def first_header_row(path: Path) -> list[str]:
    """Read only the questionnaire header row; never load participant rows."""
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style.*")
        wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        # Credamo exports can advertise an incorrect A1 used-range dimension;
        # request a bounded header span explicitly instead of trusting it.
        values = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=100, values_only=True))
        return ["" if v is None else str(v) for v in values]
    finally:
        wb.close()


def verify_raw_header_integrity(controlled: list[dict[str, Any]]) -> dict[str, Any]:
    abc_checks: dict[str, Any] = {}
    for form, instrument_path in FORM_PATHS.items():
        headers = first_header_row(SOURCE / f"{form}.xlsx")
        materials = parse_public_materials(instrument_path)
        checks = []
        for material in materials:
            exposure = normalize(material["exposure_text"])
            question = normalize(material["comprehension_question"])
            matching = [h for h in headers if exposure in normalize(h) and question in normalize(h)]
            checks.append(
                {
                    "material_no": material["material_no"],
                    "exposure_and_question_present_in_same_header": bool(matching),
                }
            )
        abc_checks[form] = {
            "header_columns": sum(bool(h) for h in headers),
            "materials_checked": len(checks),
            "all_present": all(x["exposure_and_question_present_in_same_header"] for x in checks),
            "checks": checks,
        }

    difference_headers = first_header_row(SOURCE / "差异.xlsx")
    difference_checks = []
    for case in controlled:
        question = normalize(case["question"])
        a = normalize_without_format_markers(case["version_A"])
        b = normalize_without_format_markers(case["version_B"])
        matching = [
            h
            for h in difference_headers
            if question in normalize_without_format_markers(h)
            and a in normalize_without_format_markers(h)
            and b in normalize_without_format_markers(h)
        ]
        difference_checks.append(
            {
                "case_id": case["case_id"],
                "question_and_both_versions_present_in_same_header": bool(matching),
            }
        )
    result = {
        "workbook_header_rows_read": 4,
        "participant_rows_read": False,
        "abc": abc_checks,
        "difference": {
            "header_columns": sum(bool(h) for h in difference_headers),
            "cases_checked": len(difference_checks),
            "all_present": all(x["question_and_both_versions_present_in_same_header"] for x in difference_checks),
            "checks": difference_checks,
        },
    }
    result["all_present"] = all(x["all_present"] for x in abc_checks.values()) and result["difference"]["all_present"]
    return result


def retrieve(
    query: str,
    concept_id: str | None,
    items: pd.DataFrame,
    vectorizer: TfidfVectorizer,
    matrix: Any,
    top_n: int = 5,
) -> list[dict[str, Any]]:
    if concept_id is None:
        return []
    q = vectorizer.transform([normalize(query)])
    scores = (matrix @ q.T).toarray().ravel()
    eligible = np.flatnonzero(items["concept_id"].to_numpy() == concept_id)
    order = eligible[np.argsort(scores[eligible])[::-1]][:top_n]
    return [
        {
            "rank": rank,
            "item_id": str(items.iloc[idx]["item_id"]),
            "generator": str(items.iloc[idx]["generator"]),
            "model": str(items.iloc[idx]["model"]),
            "task": str(items.iloc[idx]["task"]),
            "domain": str(items.iloc[idx]["domain"]),
            "concept_id": str(items.iloc[idx]["concept_id"]),
            "concept": str(items.iloc[idx]["concept"]),
            "score": float(scores[idx]),
        }
        for rank, idx in enumerate(order, start=1)
    ]


def row_base() -> dict[str, Any]:
    return {
        "layer": "",
        "instrument": "",
        "form_or_case": "",
        "material_no": "",
        "variant": "",
        "question_code": "",
        "selection_role": "",
        "participant_assignment": "",
        "map_status": "",
        "confidence": "",
        "mapped_item_id": "",
        "legacy_item_id": "",
        "external_exact_item_id": "",
        "generator": "",
        "model": "",
        "task": "",
        "domain": "",
        "concept_id": "",
        "concept": "",
        "core_claim": "",
        "target_misconception": "",
        "boundary_conditions": "",
        "adaptation_or_intervention": "",
        "lay_or_controlled_text": "",
        "comprehension_or_pair_question": "",
        "upstream_810_api_scores_available": "",
        "exact_text_controlled_api_scores_available": "",
        "direct_api_lay_comparison": "",
        "top_candidate_item_id": "",
        "top_candidate_score": "",
        "second_candidate_item_id": "",
        "candidate_margin": "",
        "confirmed_item_retrieval_rank": "",
    }


def main() -> None:
    ANALYSIS.mkdir(parents=True, exist_ok=True)
    NOTES.mkdir(parents=True, exist_ok=True)

    items = pd.read_csv(ITEMS_PATH)
    if len(items) != 810 or items["item_id"].nunique() != 810:
        raise RuntimeError("generated_items_810.csv failed 810-row/item-id uniqueness check")
    items = items.copy()
    items["_doc"] = items.apply(item_document, axis=1)
    vectorizer = TfidfVectorizer(analyzer="char", ngram_range=(2, 5), sublinear_tf=True, norm="l2")
    matrix = vectorizer.fit_transform(items["_doc"])

    cases = pd.read_csv(CASE_PATH)
    if set(cases["case_id"]) != {f"C{i}" for i in range(1, 7)}:
        raise RuntimeError("Expected released C1-C6 provenance bridge")
    cases_by_id = cases.set_index("case_id").to_dict(orient="index")
    item_by_id = items.set_index("item_id").to_dict(orient="index")

    controlled = json.loads(CONTROLLED_PATH.read_text(encoding="utf-8"))
    if [x["case_id"] for x in controlled] != [f"A{i:02d}" for i in range(1, 9)]:
        raise RuntimeError("Controlled contrast definition is not A01-A08")
    raw_header_integrity = verify_raw_header_integrity(controlled)
    if not raw_header_integrity["all_present"]:
        raise RuntimeError(
            "Frozen instrument text does not match raw workbook questionnaire headers: "
            + json.dumps(raw_header_integrity, ensure_ascii=False)
        )

    exact_score_ids: set[str] = set()
    exact_score_rows = 0
    exact_score_judges = 0
    if CONTROLLED_SCORE_PATH.exists():
        exact_scores = pd.read_csv(CONTROLLED_SCORE_PATH)
        exact_score_ids = set(exact_scores["item_id"].astype(str))
        exact_score_rows = len(exact_scores)
        exact_score_judges = int(exact_scores["paper_model_label"].nunique())

    rows: list[dict[str, Any]] = []
    candidate_rows: list[dict[str, Any]] = []

    # A/B/C: explicit C1-C6 -> original_item_uid bridge is decisive.
    for form, path in FORM_PATHS.items():
        for material in parse_public_materials(path):
            no = material["material_no"]
            meta = ABC_META[no]
            prov = cases_by_id[meta["source_case_id"]]
            mapped = str(prov["original_item_uid"])
            item = item_by_id[mapped]
            query = f"{material['exposure_text']} {material['comprehension_question']}"
            candidates = retrieve(query, str(item["concept_id"]), items, vectorizer, matrix)
            confirmed_rank = next((c["rank"] for c in candidates if c["item_id"] == mapped), ">5")
            base = row_base()
            base.update(
                {
                    "layer": "public_reader_ABC",
                    "instrument": f"public_reader_{form}",
                    "form_or_case": form,
                    "material_no": no,
                    "variant": form,
                    "question_code": material["qcode"],
                    "selection_role": "six released API-human disagreement examples (C1-C6); not a probability sample of 810",
                    "participant_assignment": "author-reported complete random assignment among three separate Jianshu forms",
                    "map_status": "provenance_confirmed_one_to_one",
                    "confidence": "high",
                    "mapped_item_id": mapped,
                    "legacy_item_id": str(prov["item_id"]),
                    "generator": str(item["generator"]),
                    "model": str(item["model"]),
                    "task": str(item["task"]),
                    "domain": str(item["domain"]),
                    "concept_id": str(item["concept_id"]),
                    "concept": str(item["concept"]),
                    "core_claim": meta["core_claim"],
                    "target_misconception": meta["target_misconception"],
                    "boundary_conditions": meta["boundary_conditions"],
                    "adaptation_or_intervention": "reader-facing rewrite of the same upstream item; form labels A/B/C are not corpus task labels",
                    "lay_or_controlled_text": material["exposure_text"],
                    "comprehension_or_pair_question": material["comprehension_question"],
                    "upstream_810_api_scores_available": "yes",
                    "exact_text_controlled_api_scores_available": "no",
                    "direct_api_lay_comparison": "no; adaptation changes wording, detail, boundary conditions, and sometimes repairs the source",
                    "top_candidate_item_id": candidates[0]["item_id"],
                    "top_candidate_score": candidates[0]["score"],
                    "second_candidate_item_id": candidates[1]["item_id"],
                    "candidate_margin": candidates[0]["score"] - candidates[1]["score"],
                    "confirmed_item_retrieval_rank": confirmed_rank,
                }
            )
            rows.append(base)
            for c in candidates:
                candidate_rows.append(
                    {
                        "layer": "public_reader_ABC",
                        "row_key": f"ABC_{form}_{no}",
                        "query_concept_id": item["concept_id"],
                        "confirmed_item_id": mapped,
                        "mapping_evidence": "explicit_case_provenance",
                        **c,
                    }
                )

    # Difference instrument: exact controlled text has its own API-scored IDs.
    # Concept links to 810 are valid, but a unique 810 UID is not documented.
    for case in controlled:
        case_id = case["case_id"]
        cid = CONTROLLED_CONCEPT[case_id]
        concept_rows = items.loc[items["concept_id"] == cid] if cid else items.iloc[0:0]
        domain = str(case["domain"])
        concept = str(case["concept"])
        if cid and not concept_rows.empty:
            domain = str(concept_rows.iloc[0]["domain"])
            concept = str(concept_rows.iloc[0]["concept"])
        for variant in ("A", "B"):
            text = str(case[f"version_{variant}"])
            query = f"{case['question']} {text}"
            candidates = retrieve(query, cid, items, vectorizer, matrix)
            exact_id = f"{case_id}_{variant}"
            base = row_base()
            base.update(
                {
                    "layer": "constructed_controlled_contrast",
                    "instrument": "difference",
                    "form_or_case": case_id,
                    "material_no": int(case_id[1:]),
                    "variant": variant,
                    "question_code": f"Q{5 + (int(case_id[1:]) - 1) * 3}",
                    "selection_role": "author-constructed controlled contrast with an author-specified expected direction; no probability-sampling claim",
                    "participant_assignment": "all respondents saw all eight fixed-order pairs",
                    "map_status": "outside_810_concept_frame" if cid is None else "concept_confirmed_unique_810_item_unresolved",
                    "confidence": "high_for_no_810_mapping" if cid is None else "high_for_concept_only_low_for_unique_item",
                    "mapped_item_id": "",
                    "external_exact_item_id": exact_id,
                    "domain": domain,
                    "concept_id": cid or "",
                    "concept": concept,
                    "core_claim": text,
                    "target_misconception": str(case["question"]),
                    "boundary_conditions": str(case["hidden_intervention"]),
                    "adaptation_or_intervention": f"constructed {case['hidden_intervention']} contrast; expected quality={case['expected_quality_preference']}, expected riskier={case['expected_riskier_version']}",
                    "lay_or_controlled_text": text,
                    "comprehension_or_pair_question": str(case["question"]),
                    "upstream_810_api_scores_available": "not_applicable" if cid is None else "concept-level candidates only",
                    "exact_text_controlled_api_scores_available": "yes" if exact_id in exact_score_ids else "no",
                    "direct_api_lay_comparison": "yes, using exact controlled-pair API scores; do not substitute 810-item scores",
                    "top_candidate_item_id": candidates[0]["item_id"] if candidates else "",
                    "top_candidate_score": candidates[0]["score"] if candidates else "",
                    "second_candidate_item_id": candidates[1]["item_id"] if len(candidates) > 1 else "",
                    "candidate_margin": candidates[0]["score"] - candidates[1]["score"] if len(candidates) > 1 else "",
                }
            )
            rows.append(base)
            for c in candidates:
                candidate_rows.append(
                    {
                        "layer": "constructed_controlled_contrast",
                        "row_key": f"DIFF_{case_id}_{variant}",
                        "query_concept_id": cid,
                        "confirmed_item_id": "",
                        "mapping_evidence": "retrieval_candidate_only_no_uid_provenance",
                        **c,
                    }
                )

    crosswalk = pd.DataFrame(rows)
    candidates_df = pd.DataFrame(candidate_rows)
    if len(crosswalk) != 34:
        raise RuntimeError(f"Expected 34 crosswalk rows, found {len(crosswalk)}")
    if crosswalk["mapped_item_id"].replace("", np.nan).dropna().nunique() != 6:
        raise RuntimeError("Expected six unique, explicitly confirmed 810 item IDs")

    crosswalk_path = ANALYSIS / "public_semantic_crosswalk.csv"
    candidates_path = ANALYSIS / "public_semantic_crosswalk_candidates.csv"
    audit_path = ANALYSIS / "public_semantic_crosswalk_audit.json"
    note_path = NOTES / "public_semantic_crosswalk_audit.md"
    crosswalk.to_csv(crosswalk_path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)
    candidates_df.to_csv(candidates_path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)

    input_paths = {
        "raw_A_workbook": SOURCE / "A.xlsx",
        "raw_B_workbook": SOURCE / "B.xlsx",
        "raw_C_workbook": SOURCE / "C.xlsx",
        "raw_difference_workbook": SOURCE / "差异.xlsx",
        "generated_items_810": ITEMS_PATH,
        "case_disagreement_examples": CASE_PATH,
        "controlled_pairs": CONTROLLED_PATH,
        "controlled_exact_api_scores": CONTROLLED_SCORE_PATH,
        **{f"public_instrument_{k}": v for k, v in FORM_PATHS.items()},
    }
    input_manifest = {
        k: {
            "path": str(v.relative_to(ROOT)).replace("\\", "/"),
            "bytes": v.stat().st_size,
            "sha256": sha256(v),
        }
        for k, v in input_paths.items()
        if v.exists()
    }

    abc = crosswalk[crosswalk["layer"] == "public_reader_ABC"]
    diff = crosswalk[crosswalk["layer"] == "constructed_controlled_contrast"]
    audit = {
        "schema_version": "1.0.0",
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "privacy": {
            "participant_rows_read": False,
            "participant_identifiers_written": False,
            "source_workbooks_used_for_hash_and_header_integrity_only": True,
            "output_contains_only_instrument_text_and_scientific_item_ids": True,
        },
        "raw_workbook_header_integrity": raw_header_integrity,
        "design_boundary": {
            "abc_participant_assignment": "Author reports complete random assignment among three separate Jianshu forms.",
            "abc_case_selection": "The six cases are the released C1-C6 API-human disagreement examples; no probability sampling of the 810 corpus is asserted.",
            "controlled_case_selection": "A01-A08 are constructed controlled contrasts with author-specified expected directions; no probability-sampling claim is asserted.",
            "critical_distinction": "Participant randomization does not convert purposive case selection into random item sampling.",
        },
        "retrieval_method": {
            "corpus": "810 unique generated items",
            "document_fields": ["concept", "output", "text", "all scalar content_json values"],
            "normalization": "lowercase; whitespace and common punctuation removed",
            "vectorizer": "character TF-IDF, ngram_range=(2,5), sublinear_tf=True, L2 normalization",
            "candidate_filter": "manually verified concept_id, then top five cosine similarities among 27 within-concept items",
            "decision_rule": "retrieval never creates a one-to-one mapping; only the explicit C1-C6 original_item_uid bridge does",
        },
        "counts": {
            "crosswalk_rows": int(len(crosswalk)),
            "abc_rows": int(len(abc)),
            "abc_unique_material_positions": int(abc["material_no"].nunique()),
            "abc_unique_scientific_concept_ids": int(abc["concept_id"].nunique()),
            "abc_unique_confirmed_810_items": int(abc["mapped_item_id"].nunique()),
            "controlled_rows": int(len(diff)),
            "controlled_cases": int(diff["form_or_case"].nunique()),
            "controlled_concept_links_to_810": int(diff.loc[diff["concept_id"] != "", "form_or_case"].nunique()),
            "controlled_cases_outside_810": int(diff.loc[diff["concept_id"] == "", "form_or_case"].nunique()),
            "exact_controlled_api_score_rows": exact_score_rows,
            "exact_controlled_api_judges": exact_score_judges,
            "exact_controlled_api_item_ids": len(exact_score_ids),
        },
        "abc_confirmed_sources": [
            {
                "material_no": n,
                "case_id": ABC_META[n]["source_case_id"],
                "legacy_item_id": str(cases_by_id[ABC_META[n]["source_case_id"]]["item_id"]),
                "item_id": str(cases_by_id[ABC_META[n]["source_case_id"]]["original_item_uid"]),
            }
            for n in range(1, 7)
        ],
        "controlled_810_boundary": {
            "A01": "concept C04 only; unique 810 UID unresolved",
            "A02": "concept C02 only; unique 810 UID unresolved",
            "A03": "concept C14 only; unique 810 UID unresolved",
            "A04": "concept C18 only; unique 810 UID unresolved",
            "A05": "concept C20 only; unique 810 UID unresolved",
            "A06": "concept C26 only; unique 810 UID unresolved",
            "A07": "Chemical equilibrium is absent from the 30-concept 810 frame; exact external controlled IDs A07_A/A07_B exist",
            "A08": "concept C25 only; unique 810 UID unresolved",
        },
        "comparison_validity": {
            "abc": {
                "direct_original_api_vs_lay_outcome": False,
                "reason": "The public text was rewritten and differs in length, mechanism detail, boundary conditions, and sometimes error repair from the upstream 810 item.",
                "defensible_role": "randomized reader-response experiment over three adaptations of six traced disagreement cases; upstream item ID is provenance, not an unchanged stimulus",
                "recommended_model": "participant-concept post_correct ~ pre_correct + form + concept, using a binomial GEE clustered by participant or a mixed logistic model with participant intercept; form-by-concept interaction exploratory",
                "generalization_limit": "six purposive disagreement cases and one text realization per form-concept; do not generalize a form effect to the full 810 corpus",
            },
            "controlled": {
                "direct_810_score_comparison": False,
                "exact_controlled_text_comparison": True,
                "reason": "All 16 displayed texts have exact, separately scored controlled IDs A01_A-A08_B; those scores match the stimuli, whereas 810 candidates do not.",
                "recommended_human_model": "expected-direction forced choice with crossed participant and case intercepts, supplemented by case-level exact binomial intervals and multiplicity control",
                "recommended_api_human_model": "for each case, compare human expected-choice share with exact-text judge score contrasts; use an eight-case exact-permutation Spearman analysis and report case-level heterogeneity",
                "generalization_limit": "engineered contrasts test discrimination/mechanism salience, not natural-error prevalence or corpus-wide validity",
            },
        },
        "outputs": {
            "crosswalk_csv": str(crosswalk_path.relative_to(ROOT)).replace("\\", "/"),
            "candidate_csv": str(candidates_path.relative_to(ROOT)).replace("\\", "/"),
            "audit_json": str(audit_path.relative_to(ROOT)).replace("\\", "/"),
            "audit_markdown": str(note_path.relative_to(ROOT)).replace("\\", "/"),
        },
        "inputs": input_manifest,
    }
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    serialized_audit = audit_path.read_text(encoding="utf-8")
    mojibake_markers = ["鏁", "閲", "程", "鍏", "鍙"]
    if any(marker in serialized_audit for marker in mojibake_markers):
        raise RuntimeError("Mojibake marker found in audit JSON path/text fields")

    source_lines = "\n".join(
        f"| {n} | {ABC_META[n]['source_case_id']} | {cases_by_id[ABC_META[n]['source_case_id']]['original_item_uid']} | "
        f"{cases_by_id[ABC_META[n]['source_case_id']]['generator']} | {cases_by_id[ABC_META[n]['source_case_id']]['task']} | "
        f"{cases_by_id[ABC_META[n]['source_case_id']]['concept']} |"
        for n in range(1, 7)
    )
    md = f"""# Public-material semantic/provenance crosswalk audit

## Bottom line

The 18 A/B/C exposures are **three reader-facing adaptations of the same six upstream disagreement cases**, not 18 independent 810-item draws.  The released `case_disagreement_examples.csv` provides an explicit one-to-one bridge through `original_item_uid`, so these six source mappings are high-confidence provenance mappings.  The author reports that respondents were completely randomly assigned among the three separate Jianshu forms; that participant randomization is distinct from item selection.

The 16 texts in `差异.xlsx` are the exact A01-A08 constructed controlled contrasts.  Seven cases have a clear 810 **concept** link, but there is no surviving unique-UID bridge and the texts contain deliberate interventions; no individual 810 item is therefore asserted as their source.  A07 chemical equilibrium is outside the current 30-concept/810-item frame.  It is nevertheless fully traceable in the separate exact-text controlled API dataset as `A07_A` and `A07_B`.

## Confirmed A/B/C upstream sources

| Material | Released case | 810 item ID | Generator | Corpus task | Concept |
|---:|---|---|---|---|---|
{source_lines}

Every A, B, and C version of a numbered material maps to the same row above.  The A/B/C form label must not be interpreted as the corpus task label.

## Difference/controlled-contrast boundary

| Case | 810 concept link | Unique 810 item | Exact controlled API IDs |
|---|---|---|---|
| A01 | C04 Projectile motion | unresolved | A01_A, A01_B |
| A02 | C02 Net force and acceleration | unresolved | A02_A, A02_B |
| A03 | C14 Evolution has no predetermined goal | unresolved | A03_A, A03_B |
| A04 | C18 Photosynthesis and cellular respiration | unresolved | A04_A, A04_B |
| A05 | C20 Geological time | unresolved | A05_A, A05_B |
| A06 | C26 Weather and climate | unresolved | A06_A, A06_B |
| A07 | absent from the 810 concept frame | none | A07_A, A07_B |
| A08 | C25 Greenhouse effect | unresolved | A08_A, A08_B |

The candidate CSV records reproducible character TF-IDF retrieval within each manually verified concept.  Candidate ranking is diagnostic: semantic similarity alone cannot identify which of 27 same-concept corpus items, if any, was the historical anchor after author construction.

## Can the API scores be compared with reader outcomes?

**A/B/C:** not as a direct score-validity test.  The upstream six IDs are recoverable, but the public stimuli were rewritten for lay comprehension and differ in length, mechanism detail, boundary conditions, and sometimes error repair.  Original 810 API scores may be used only as upstream case context.  A defensible analysis treats A/B/C as a randomized reader-response experiment: model post-reading correctness from pre-reading correctness, form, and concept with participant-clustered GEE or a participant-intercept mixed logistic model.  Because the six cases are disagreement examples rather than a probability sample, any form effect is local to these cases.

**Difference:** yes, but only with the separately stored exact-text controlled API scores, not with forced 810 mappings.  Each displayed text has an exact ID `A01_A`-`A08_B` scored by the API panel.  A defensible analysis estimates expected-direction forced choice with participant and case effects, reports each case, and relates human expected-choice share to exact-text API score contrasts across the eight cases using exact permutation inference.  These engineered contrasts test discriminative sensitivity/mechanism salience; they do not estimate natural-error prevalence.

## Design and evidential boundaries

- Author-reported randomization applies to **participants across A/B/C forms**.
- The six A/B/C cases are released C1-C6 API-human disagreement examples; no probability sampling of 810 is claimed.
- A01-A08 are constructed contrasts with author-specified expected directions; no probability-sampling claim is made.
- Participant randomization and item sampling are different design properties.
- No participant row or private identifier was read into or written by this crosswalk pipeline.

## Reproducibility

The JSON audit records SHA-256 hashes of the four source workbooks, frozen privacy-safe instruments, the 810 corpus, the explicit C1-C6 bridge, the A01-A08 definition, and the exact controlled API score file.  The crosswalk has {len(crosswalk)} rows (18 A/B/C + 16 controlled texts); only six unique 810 IDs are labelled one-to-one, all through explicit provenance.
"""
    note_path.write_text(md, encoding="utf-8")

    # Re-open deliverables for compact QA.
    check_crosswalk = pd.read_csv(crosswalk_path)
    check_candidates = pd.read_csv(candidates_path)
    check_audit = json.loads(audit_path.read_text(encoding="utf-8"))
    assert len(check_crosswalk) == 34
    assert len(check_candidates) == len(candidate_rows)
    assert check_audit["counts"]["abc_unique_confirmed_810_items"] == 6
    privacy_fields = {
        "response_id",
        "user_id",
        "ip",
        "longitude",
        "latitude",
        "device",
        "timestamp",
        "start_time",
        "end_time",
    }
    assert privacy_fields.isdisjoint({str(c).lower() for c in check_crosswalk.columns})
    print(
        json.dumps(
            {
                "status": "PASS",
                "crosswalk_rows": len(check_crosswalk),
                "candidate_rows": len(check_candidates),
                "confirmed_810_ids": int(
                    check_crosswalk["mapped_item_id"].dropna().replace("", np.nan).dropna().nunique()
                ),
                "outside_810_cases": sorted(
                    check_crosswalk.loc[
                        check_crosswalk["map_status"] == "outside_810_concept_frame", "form_or_case"
                    ].unique()
                ),
                "outputs": [str(crosswalk_path), str(candidates_path), str(audit_path), str(note_path)],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
